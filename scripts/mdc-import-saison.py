#!/usr/bin/env python3
# ============================================================
# MDC — eine Saison aus der Arbeitsmappe des Betreibers einlesen
# ============================================================
#
# Quelle ist die Excel-Arbeitsmappe, mit der die MDC ihre Saison führt
# („MDC_2025_2026.xlsm", „MDC_2026_2027.xlsm"). Sie liegt bewusst NICHT im
# Repository: rund 9 MB und darin das komplette Teilnehmerregister mit allen
# Namen. Ins Repository kommen nur die Ergebnisse, die ohnehin veröffentlicht
# werden.
#
#   Aufruf:  python3 scripts/mdc-import-saison.py <mappe.xlsm> <saison>
#   Saison:  2025-26 | 2026-27
#   Braucht: pip install openpyxl
#
# Gelesen werden drei Blätter:
#
#   „Einzelergebnisse"  Alle Turniere der Saison, in Blöcken untereinander.
#                       Spalte E: erst der Spielort, in der Zeile darunter das
#                       Datum. Spalte F trägt je Zeile den Blockschlüssel
#                       (Spielort + Excel-Datumszahl), G–M die Ergebniszeile:
#                       Platz, Passnr., Name, Vorname, weibl., Teilnehmer,
#                       Punkte.
#   „Männer"/„Frauen"   Die Rangliste zum Stand der Mappe.
#
# Erzeugt (jeweils komplett überschrieben):
#
#   data/results-<saison>.generated.ts
#   data/ranking-<saison>-men.ts
#   data/ranking-<saison>-women.ts
#
# Das Skript prüft dabei, was sich prüfen lässt, und bricht bei Widersprüchen
# ab — lieber kein Import als ein stiller Fehler:
#
#   • Jeder Block: Platzfolge 1..n lückenlos, Zeilenzahl = Teilnehmerzahl,
#     Teilnehmerzahl in allen Zeilen gleich.
#   • Punkte je (Feldgröße, Platz) über alle Turniere hinweg widerspruchsfrei.
#   • Summe der Einzelpunkte je Passnummer = Punkte der Rangliste,
#     Anzahl der Starts = Spalte „Anzahl TN".
#   • Name je Passnummer in allen Ergebniszeilen gleich.
#   • Jede Ergebniszeile hat einen Ranglistenplatz.
# ============================================================

import re
import sys
import datetime
from collections import defaultdict, OrderedDict

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit('Fehlt: openpyxl (pip install openpyxl)')

# Spielort-Bezeichnung in der Mappe → ID in data/venues.ts. Die Mappe schreibt
# manche Lokale unterschiedlich („5 STERNE BOAZN" / „5STERNE BOAZN"), deshalb
# stehen mehrere Schreibweisen nebeneinander. Spielorte, in denen heute nicht
# mehr gespielt wird, stehen als `FORMER_VENUES` in `data/venues.ts` — nur mit
# Namen, ohne erfundene Adresse.
VENUE_IDS = {
    'TONYS': 'tonys-wirtshaus',
    'LEGENDARY': 'legendary',
    'HARLEKIN': 'harlekin',
    'BISTRO 118': 'bistro-118',
    'AMBASADOR': 'ambasador',
    'WÜRMTAL': 'djk-wuermtal',
    '5STERNE BOAZN': 'fuenf-sterne-boazn',
    '5 STERNE BOAZN': 'fuenf-sterne-boazn',
    'MACHETE': 'machete-1',
    '70ER': 'siebziger',
    'LUSTIGER BAUER': 'lustiger-bauer',
    'FIAKER': 'fiakerstueberl',
    'RG BAR': 'rg-bar',
    'GISI': 'gisi',
    'FLAIR': 'flair',
}

SAISONS = {
    '2025-26': {
        'const': '2025_26',
        'label': '2025/26',
        # Die Männerwertung führt einen Lückenmarker — sie wurde einmal aus
        # unvollständigen Unterlagen aufgebaut (siehe docs/mdc-demo.md).
        'gap_marker': True,
        'herkunft': (
            '// Vor dem ersten Import (September 2026) stand hier eine von Fotos der\n'
            '// gedruckten Auswertung abgetippte Fassung. Der Abgleich mit der Mappe hat\n'
            '// 13 Lesefehler in der Männer- und 2 in der Frauenwertung berichtigt, sieben\n'
            '// übersehene geteilte Plätze erkannt und eine fehlende Zeile ergänzt.\n'
        ),
    },
    '2026-27': {
        'const': '2026_27',
        'label': '2026/27',
        'gap_marker': False,
        'herkunft': (
            '// Die laufende Saison: Diese Datei wird bei jeder neuen Fassung der Mappe\n'
            '// neu erzeugt. Vorher standen hier acht von Hand abgetippte Ergebniszettel;\n'
            '// die Mappe hat sie ersetzt (siehe docs/mdc-demo.md).\n'
        ),
    },
}


def zelle(v):
    if v is None:
        return ''
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime('%d.%m.%Y')
    return str(v).strip()


def namen_normalisieren(text):
    """„MARCUS ( HECHI )" und „CHRISTOPH(STOFFEL)" → „MARCUS (HECHI)"."""
    text = re.sub(r'\s*\(\s*', ' (', text)
    text = re.sub(r'\s*\)', ')', text)
    return re.sub(r'\s+', ' ', text).strip()


def lies_turniere(ws):
    turniere = OrderedDict()
    for row in ws.iter_rows(min_row=10, max_col=13, values_only=True):
        ort, key, platz, passnr, name, vorname, weibl, tn, punkte = (
            zelle(row[i]) for i in range(4, 13)
        )
        if not key:
            if any([platz, passnr, name, punkte]):
                sys.exit(f'Ergebniszeile ohne Blockschlüssel: {row}')
            continue
        block = turniere.setdefault(key, {'ort': '', 'datum': '', 'rows': []})
        if ort:
            if re.match(r'^\d{2}\.\d{2}\.\d{4}$', ort):
                block['datum'] = ort
            else:
                block['ort'] = ort
        if platz:
            block['rows'].append({
                'platz': int(float(platz)),
                'pass': int(float(passnr)),
                'name': namen_normalisieren(name),
                'vorname': namen_normalisieren(vorname),
                'weibl': bool(weibl),
                'tn': int(float(tn)),
                'punkte': int(float(punkte)),
            })
    return turniere


def lies_stand(ws):
    """Datum aus der Kopfzeile des Ranglistenblatts („… vom 05.09.2026")."""
    for row in ws.iter_rows(min_row=1, max_row=1, max_col=20, values_only=True):
        for wert in row:
            if isinstance(wert, (datetime.datetime, datetime.date)):
                return wert.strftime('%Y-%m-%d')
    return ''


def lies_rangliste(ws):
    """Spalten der Auswertung: H Platz, I Trend, J Passnr., K Name, L Vorname,
    M Anzahl TN, N Punkte, Q %."""
    zeilen = []
    for row in ws.iter_rows(min_row=3, max_col=26, values_only=True):
        if row[7] is None or row[9] is None:
            continue
        zeilen.append({
            'platz': int(row[7]),
            'trend': {'▲': 'u', '▼': 'd'}.get(zelle(row[8]), ''),
            'pass': int(row[9]),
            'name': namen_normalisieren(zelle(row[10])),
            'vorname': namen_normalisieren(zelle(row[11])),
            'starts': int(row[12]),
            'punkte': int(row[13]),
            'prozent': row[16],
        })
    return zeilen


def pruefe_turniere(turniere):
    punkte_je_feld = defaultdict(dict)
    for key, t in turniere.items():
        if not t['ort'] or not t['datum']:
            sys.exit(f'Block ohne Spielort oder Datum: {key}')
        tn = {r['tn'] for r in t['rows']}
        if len(tn) != 1:
            sys.exit(f'{key}: uneinheitliche Teilnehmerzahl {tn}')
        tn = tn.pop()
        if len(t['rows']) != tn:
            sys.exit(f'{key}: {len(t["rows"])} Zeilen, aber {tn} Teilnehmer')
        if [r['platz'] for r in t['rows']] != list(range(1, tn + 1)):
            sys.exit(f'{key}: Plätze nicht lückenlos 1..{tn}')
        for r in t['rows']:
            bekannt = punkte_je_feld[tn].get(r['platz'])
            if bekannt is not None and bekannt != r['punkte']:
                sys.exit(f'{key}: {tn} TN, Platz {r["platz"]} → {r["punkte"]}, '
                         f'anderswo {bekannt}')
            punkte_je_feld[tn][r['platz']] = r['punkte']
    return punkte_je_feld


def pruefe_abgleich(turniere, maenner, frauen):
    summe = defaultdict(int)
    starts = defaultdict(int)
    namen = {}
    for t in turniere.values():
        for r in t['rows']:
            summe[r['pass']] += r['punkte']
            starts[r['pass']] += 1
            voll = (r['name'], r['vorname'])
            if namen.setdefault(r['pass'], voll) != voll:
                sys.exit(f'Passnr. {r["pass"]} trägt zwei Namen: '
                         f'{namen[r["pass"]]} / {voll}')
    for eintrag in maenner + frauen:
        p = eintrag['pass']
        if summe[p] != eintrag['punkte'] or starts[p] != eintrag['starts']:
            sys.exit(f'Passnr. {p}: Einzelergebnisse ergeben {summe[p]} Punkte aus '
                     f'{starts[p]} Starts, die Rangliste nennt {eintrag["punkte"]} '
                     f'aus {eintrag["starts"]}')
    gewertet = {e['pass'] for e in maenner + frauen}
    ohne = sorted(set(summe) - gewertet)
    if ohne:
        sys.exit(f'Passnummern mit Ergebnissen, aber ohne Ranglistenplatz: {ohne}')
    return sum(summe.values()), sum(starts.values())


def rang_zeilen(zeilen):
    """Zeilen der Rangliste im Format der data/ranking-*.ts.

    Geteilte Plätze werden aus den Punkten abgeleitet: gleiche Punktzahl wie
    die Zeile darüber → Platz bleibt leer. Genau so ist das Format definiert.
    """
    out = []
    vorherige_punkte = None
    for z in zeilen:
        geteilt = vorherige_punkte is not None and z['punkte'] == vorherige_punkte
        prozent = '' if z['prozent'] in (None, '') else f"{round(float(z['prozent']) * 100, 4):g}"
        felder = [
            '' if geteilt else str(z['platz']),
            str(z['pass']), z['name'], z['vorname'],
            str(z['starts']), str(z['punkte']),
        ]
        if prozent or z['trend']:
            felder.append(prozent)
        if z['trend']:
            felder.append(z['trend'])
        out.append('|'.join(felder))
        vorherige_punkte = z['punkte']
    return out


def schreibe(pfad, text):
    with open(pfad, 'w', encoding='utf-8') as f:
        f.write(text)
    print(f'  geschrieben: {pfad}')


def kopf_rangliste(saison, division, anzahl, geteilt, stand):
    titel = 'Männer' if division == 'men' else 'Frauen'
    return f'''// ============================================================
// MDC — Rangliste {titel}, Saison {saison["label"]} (Stand {stand})
// ============================================================
//
// ERZEUGT aus der Arbeitsmappe des Betreibers durch
// `scripts/mdc-import-saison.py` (Blatt „{titel}"). Nicht von Hand
// bearbeiten — sonst laufen Rangliste und Einzelergebnisse auseinander.
//
// Format je Zeile:
//
//   Platz | Passnr. | Name | Vorname | Anzahl TN | Punkte | % | Trend
//
// • Platz leer  → punktgleich mit der Zeile darüber (geteilter Platz).
//                 {geteilt} der {anzahl} Zeilen teilen sich so einen Platz.
// • %           → Anteil an der Einzelranglisten-Ausschüttung (EZR).
//                 Der Euro-Betrag wird daraus berechnet (siehe ranking-final.ts),
//                 damit Prozent und Euro nicht auseinanderlaufen können.
// • Trend       → 'u' = ▲ gestiegen, 'd' = ▼ gefallen, leer = unverändert.
// • Schnitt     → wird als Punkte / Anzahl TN berechnet, nicht gepflegt.
//
// Jede Zeile ist gegen die Einzelergebnisse derselben Mappe gerechnet: Die
// Summe der Turnierpunkte einer Passnummer ergibt die Punktzahl hier, die
// Anzahl der Starts die Spalte „Anzahl TN".
//
{saison["herkunft"]}// ============================================================
'''


def main():
    if len(sys.argv) < 3 or sys.argv[2] not in SAISONS:
        sys.exit('Aufruf: mdc-import-saison.py <mappe.xlsm> <' +
                 ' | '.join(SAISONS) + '>')
    pfad, saison_id = sys.argv[1], sys.argv[2]
    saison = SAISONS[saison_id]
    wb = openpyxl.load_workbook(pfad, data_only=True, read_only=True)

    turniere = lies_turniere(wb['Einzelergebnisse'])
    stand = lies_stand(wb['Männer'])
    maenner = lies_rangliste(wb['Männer'])
    frauen = lies_rangliste(wb['Frauen'])
    print(f'gelesen: {len(turniere)} Turniere, '
          f'{sum(len(t["rows"]) for t in turniere.values())} Ergebniszeilen, '
          f'{len(maenner)} + {len(frauen)} Ranglistenzeilen')

    pruefe_turniere(turniere)
    punkte, starts = pruefe_abgleich(turniere, maenner, frauen)
    print(f'geprüft: {starts} Starts, {punkte} Punkte — Einzelergebnisse und '
          f'Rangliste stimmen überein')

    # ---- Einzelergebnisse ----
    zeilen = []
    unbekannt = set()
    for t in turniere.values():
        vid = VENUE_IDS.get(t['ort'].upper())
        if not vid:
            unbekannt.add(t['ort'])
            continue
        tag = datetime.datetime.strptime(t['datum'], '%d.%m.%Y').strftime('%Y-%m-%d')
        ergebnisse = ','.join(f"{r['pass']}:{r['punkte']}" for r in t['rows'])
        zeilen.append((tag, vid, f'{tag}|{vid}|{ergebnisse}'))
    if unbekannt:
        sys.exit(f'Spielort ohne ID in VENUE_IDS: {sorted(unbekannt)}')
    zeilen.sort(key=lambda z: (z[0], z[1]))
    # Ohne Datum in der Kopfzeile gilt der letzte Turniertag als Stand.
    stand = stand or max(z[0] for z in zeilen)
    print(f'Stand der Rangliste: {stand}')

    body = ''.join(f"  '{z[2]}',\n" for z in zeilen)
    schreibe(f'data/results-{saison_id}.generated.ts', f'''// ============================================================
// MDC — Einzelergebnisse der Saison {saison["label"]}
// ============================================================
//
// ERZEUGT aus der Arbeitsmappe des Betreibers durch
// `scripts/mdc-import-saison.py` (Blatt „Einzelergebnisse").
// Nicht von Hand bearbeiten.
//
// {len(zeilen)} Turniere vom {zeilen[0][0]} bis {zeilen[-1][0]}, zusammen
// {starts} Ergebniszeilen und {punkte} vergebene Punkte.
//
// Eine Zeile je Turnier:
//
//   Datum | Spielort-ID | Passnr.:Punkte, Passnr.:Punkte, …
//
// Die Ergebnisse stehen in Platzreihenfolge — der erste Eintrag ist Platz 1.
// Die Teilnehmerzahl ist die Anzahl der Einträge; jedes Turnier der Mappe hat
// genau so viele Zeilen wie Teilnehmer und eine lückenlose Platzfolge.
//
// Die Punkte stehen mit dabei, obwohl sie sich aus Platz und Feldgröße
// ergeben würden (`lib/mdc/points.ts`): Sie sind das, was der Betreiber
// tatsächlich verbucht hat. `scripts/mdc-check-saison.ts` rechnet beides
// gegeneinander.
// ============================================================

export const RESULTS_{saison["const"]}_RAW: string[] = [
{body}];
''')

    # ---- Ranglisten ----
    for division, zeilen_r, datei, const in [
        ('men', maenner, f'data/ranking-{saison_id}-men.ts',
         f'RANKING_MEN_{saison["const"]}_RAW'),
        ('women', frauen, f'data/ranking-{saison_id}-women.ts',
         f'RANKING_WOMEN_{saison["const"]}_RAW'),
    ]:
        rows = rang_zeilen(zeilen_r)
        geteilt = sum(1 for r in rows if r.startswith('|'))
        body = ''.join(f"  '{r}',\n" for r in rows)
        gap = ''
        if division == 'men' and saison['gap_marker']:
            gap = ('\n/**\n'
                   ' * Plätze, deren Auswertungsseiten fehlen — wird in der Tabelle\n'
                   ' * ausgewiesen. `null` = keine Lücke. Seit dem Import aus der\n'
                   ' * Arbeitsmappe ist die Wertung vollständig; die Plumbing bleibt für\n'
                   ' * den Fall, dass später einmal etwas fehlt.\n'
                   ' */\n'
                   'export const RANKING_MEN_GAP: { from: number; to: number } | null = null;\n')
        schreibe(datei,
                 kopf_rangliste(saison, division, len(rows), geteilt, stand)
                 + f'\nexport const {const}: string[] = [\n{body}];\n' + gap)


if __name__ == '__main__':
    main()

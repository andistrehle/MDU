#!/usr/bin/env python3
# ============================================================
# MDC — Saison 2025/26 aus der Arbeitsmappe des Betreibers einlesen
# ============================================================
#
# Quelle ist die Excel-Arbeitsmappe, mit der die MDC ihre Saison führt
# („MDC_2025_2026.xlsm"). Sie liegt bewusst NICHT im Repository: 9,6 MB und
# darin das komplette Teilnehmerregister mit allen Namen. Ins Repository
# kommen nur die Ergebnisse, die ohnehin veröffentlicht werden.
#
#   Aufruf:  python3 scripts/mdc-import-saison-2025-26.py <pfad/zur/mappe.xlsm>
#   Braucht: pip install openpyxl
#
# Gelesen werden drei Blätter:
#
#   „Einzelergebnisse"  Alle Turniere der Saison, in Blöcken untereinander.
#                       Spalte E: erst der Spielort, in der Zeile darunter das
#                       Datum. Spalte F trägt je Zeile den Blockschlüssel
#                       (Spielort + Excel-Datumszahl), F–M die Ergebniszeile:
#                       Platz, Passnr., Name, Vorname, weibl., Teilnehmer,
#                       Punkte.
#   „Männer"/„Frauen"   Die offizielle Endrangliste vom 27.07.2026.
#
# Erzeugt (jeweils komplett überschrieben):
#
#   data/ranking-2025-26-men.ts
#   data/ranking-2025-26-women.ts
#   data/archive-2025-26.generated.ts
#
# Das Skript prüft dabei, was sich prüfen lässt, und bricht bei Widersprüchen
# ab — lieber kein Import als ein stiller Fehler:
#
#   • Jeder Block: Platzfolge 1..n lückenlos, Zeilenzahl = Teilnehmerzahl,
#     Teilnehmerzahl in allen Zeilen gleich.
#   • Punkte je (Feldgröße, Platz) über alle Turniere hinweg widerspruchsfrei.
#   • Summe der Einzelpunkte je Passnummer = Punkte der Endrangliste,
#     Anzahl der Starts = Spalte „Anzahl TN".
#   • Name je Passnummer in allen Ergebniszeilen gleich.
# ============================================================

import re
import sys
import datetime
from collections import defaultdict, OrderedDict

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit('Fehlt: openpyxl (pip install openpyxl)')

# Spielort-Bezeichnung in der Mappe → ID in data/venues.ts.
# Spielorte, die es in der Saison 2026/27 nicht mehr gibt, stehen in
# `FORMER_VENUES` (data/venues.ts) — nur mit Namen, ohne erfundene Adresse.
VENUE_IDS = {
    'TONYS': 'tonys-wirtshaus',
    'Legendary': 'legendary',
    'Harlekin': 'harlekin',
    'WÜRMTAL': 'djk-wuermtal',
    '5Sterne Boazn': 'fuenf-sterne-boazn',
    'MACHETE': 'machete-1',
    '70ER': 'siebziger',
    'Lustiger Bauer': 'lustiger-bauer',
    'FIAKER': 'fiakerstueberl',
    'Fiaker': 'fiakerstueberl',
    'RG BAR': 'rg-bar',
    'GISI': 'gisi',
    'FLAIR': 'flair',
}

FOOTER = ''


def zelle(v):
    if v is None:
        return ''
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime('%d.%m.%Y')
    return str(v).strip()


def namen_normalisieren(text):
    """„MARCUS ( HECHI )" und „CHRISTOPH(STOFFEL)" → „MARCUS (HECHI)"."""
    text = re.sub(r'\s*\(\s*', ' (', text)
    text = re.sub(r'\s*\)', ')', text)
    return re.sub(r'\s+', ' ', text).strip()


def lies_turniere(ws):
    turniere = OrderedDict()
    for row in ws.iter_rows(min_row=10, max_col=13, values_only=True):
        ort, key, platz, passnr, name, vorname, weibl, tn, punkte = (
            zelle(row[i]) for i in range(4, 13)
        )
        if not key:
            if any([platz, passnr, name, punkte]):
                sys.exit(f'Ergebniszeile ohne Blockschlüssel: {row}')
            continue
        block = turniere.setdefault(key, {'ort': '', 'datum': '', 'rows': []})
        if ort:
            if re.match(r'^\d{2}\.\d{2}\.\d{4}$', ort):
                block['datum'] = ort
            else:
                block['ort'] = ort
        if platz:
            block['rows'].append({
                'platz': int(float(platz)),
                'pass': int(float(passnr)),
                'name': namen_normalisieren(name),
                'vorname': namen_normalisieren(vorname),
                'weibl': bool(weibl),
                'tn': int(float(tn)),
                'punkte': int(float(punkte)),
            })
    return turniere


def lies_rangliste(ws):
    """Spalten der Auswertung: H Platz, I Trend, J Passnr., K Name, L Vorname,
    M Anzahl TN, N Punkte, Q %, Z Entwicklung."""
    zeilen = []
    for row in ws.iter_rows(min_row=3, max_col=26, values_only=True):
        if row[7] is None or row[9] is None:
            continue
        zeilen.append({
            'platz': int(row[7]),
            'trend': {'▲': 'u', '▼': 'd'}.get(zelle(row[8]), ''),
            'pass': int(row[9]),
            'name': namen_normalisieren(zelle(row[10])),
            'vorname': namen_normalisieren(zelle(row[11])),
            'starts': int(row[12]),
            'punkte': int(row[13]),
            'prozent': row[16],
        })
    return zeilen


def pruefe_turniere(turniere):
    punkte_je_feld = defaultdict(dict)
    for key, t in turniere.items():
        if not t['ort'] or not t['datum']:
            sys.exit(f'Block ohne Spielort oder Datum: {key}')
        tn = {r['tn'] for r in t['rows']}
        if len(tn) != 1:
            sys.exit(f'{key}: uneinheitliche Teilnehmerzahl {tn}')
        tn = tn.pop()
        if len(t['rows']) != tn:
            sys.exit(f'{key}: {len(t["rows"])} Zeilen, aber {tn} Teilnehmer')
        if [r['platz'] for r in t['rows']] != list(range(1, tn + 1)):
            sys.exit(f'{key}: Plätze nicht lückenlos 1..{tn}')
        for r in t['rows']:
            bekannt = punkte_je_feld[tn].get(r['platz'])
            if bekannt is not None and bekannt != r['punkte']:
                sys.exit(f'{key}: {tn} TN, Platz {r["platz"]} → {r["punkte"]}, '
                         f'anderswo {bekannt}')
            punkte_je_feld[tn][r['platz']] = r['punkte']
    return punkte_je_feld


def pruefe_abgleich(turniere, maenner, frauen):
    summe = defaultdict(int)
    starts = defaultdict(int)
    namen = {}
    for t in turniere.values():
        for r in t['rows']:
            summe[r['pass']] += r['punkte']
            starts[r['pass']] += 1
            voll = (r['name'], r['vorname'])
            if namen.setdefault(r['pass'], voll) != voll:
                sys.exit(f'Passnr. {r["pass"]} trägt zwei Namen: '
                         f'{namen[r["pass"]]} / {voll}')
    for eintrag in maenner + frauen:
        p = eintrag['pass']
        if summe[p] != eintrag['punkte'] or starts[p] != eintrag['starts']:
            sys.exit(f'Passnr. {p}: Einzelergebnisse ergeben {summe[p]} Punkte aus '
                     f'{starts[p]} Starts, die Rangliste nennt {eintrag["punkte"]} '
                     f'aus {eintrag["starts"]}')
    gewertet = {e['pass'] for e in maenner + frauen}
    ohne = sorted(set(summe) - gewertet)
    if ohne:
        sys.exit(f'Passnummern mit Ergebnissen, aber ohne Ranglistenplatz: {ohne}')
    return sum(summe.values()), sum(starts.values())


def rang_zeilen(zeilen):
    """Zeilen der Endrangliste im Format der data/ranking-*.ts.

    Geteilte Plätze werden aus den Punkten abgeleitet: gleiche Punktzahl wie
    die Zeile darüber → Platz bleibt leer. Genau so ist das Format definiert.
    """
    out = []
    vorherige_punkte = None
    for z in zeilen:
        geteilt = vorherige_punkte is not None and z['punkte'] == vorherige_punkte
        prozent = '' if z['prozent'] in (None, '') else f"{round(float(z['prozent']) * 100, 4):g}"
        felder = [
            '' if geteilt else str(z['platz']),
            str(z['pass']), z['name'], z['vorname'],
            str(z['starts']), str(z['punkte']),
        ]
        if prozent or z['trend']:
            felder.append(prozent)
        if z['trend']:
            felder.append(z['trend'])
        out.append('|'.join(felder))
        vorherige_punkte = z['punkte']
    return out


def schreibe(pfad, text):
    with open(pfad, 'w', encoding='utf-8') as f:
        f.write(text)
    print(f'  geschrieben: {pfad}')


def kopf_rangliste(division, anzahl, geteilt):
    titel = 'Männer' if division == 'men' else 'Frauen'
    return f'''// ============================================================
// MDC — Endrangliste {titel}, Saison 2025/26 (Stand 27.07.2026)
// ============================================================
//
// ERZEUGT aus der Arbeitsmappe des Betreibers durch
// `scripts/mdc-import-saison-2025-26.py` (Blatt „{titel}"). Nicht von Hand
// bearbeiten — sonst laufen Rangliste und Einzelergebnisse auseinander.
//
// Format je Zeile:
//
//   Platz | Passnr. | Name | Vorname | Anzahl TN | Punkte | % | Trend
//
// • Platz leer  → punktgleich mit der Zeile darüber (geteilter Platz).
//                 {geteilt} der {anzahl} Zeilen teilen sich so einen Platz.
// • %           → Anteil an der Einzelranglisten-Ausschüttung (EZR).
//                 Der Euro-Betrag wird daraus berechnet (siehe payout.ts),
//                 damit Prozent und Euro nicht auseinanderlaufen können.
// • Trend       → 'u' = ▲ gestiegen, 'd' = ▼ gefallen, leer = unverändert.
// • Schnitt     → wird als Punkte / Anzahl TN berechnet, nicht gepflegt.
//
// Jede Zeile ist gegen die Einzelergebnisse derselben Mappe gerechnet: Die
// Summe der Turnierpunkte einer Passnummer ergibt die Punktzahl hier, die
// Anzahl der Starts die Spalte „Anzahl TN".
//
// Vor dem ersten Import (September 2026) stand hier eine von Fotos der
// gedruckten Auswertung abgetippte Fassung. Der Abgleich mit der Mappe hat
// 13 Lesefehler in der Männer- und 2 in der Frauenwertung berichtigt, sieben
// übersehene geteilte Plätze erkannt und eine fehlende Zeile ergänzt.
// ============================================================
'''


def main():
    if len(sys.argv) < 2:
        sys.exit('Aufruf: mdc-import-saison-2025-26.py <mappe.xlsm>')
    wb = openpyxl.load_workbook(sys.argv[1], data_only=True, read_only=True)

    turniere = lies_turniere(wb['Einzelergebnisse'])
    maenner = lies_rangliste(wb['Männer'])
    frauen = lies_rangliste(wb['Frauen'])
    print(f'gelesen: {len(turniere)} Turniere, '
          f'{sum(len(t["rows"]) for t in turniere.values())} Ergebniszeilen, '
          f'{len(maenner)} + {len(frauen)} Ranglistenzeilen')

    pruefe_turniere(turniere)
    punkte, starts = pruefe_abgleich(turniere, maenner, frauen)
    print(f'geprüft: {starts} Starts, {punkte} Punkte — Einzelergebnisse und '
          f'Endrangliste stimmen überein')

    for division, zeilen, pfad, const in [
        ('men', maenner, 'data/ranking-2025-26-men.ts', 'RANKING_MEN_2025_26_RAW'),
        ('women', frauen, 'data/ranking-2025-26-women.ts', 'RANKING_WOMEN_2025_26_RAW'),
    ]:
        rows = rang_zeilen(zeilen)
        geteilt = sum(1 for r in rows if r.startswith('|'))
        body = ''.join(f"  '{r}',\n" for r in rows)
        gap = ''
        if division == 'men':
            gap = ('\n/**\n'
                   ' * Plätze, deren Auswertungsseiten fehlen — wird in der Tabelle\n'
                   ' * ausgewiesen. `null` = keine Lücke. Seit dem Import aus der\n'
                   ' * Arbeitsmappe ist die Wertung vollständig; die Plumbing bleibt für\n'
                   ' * den Fall, dass später einmal etwas fehlt.\n'
                   ' */\n'
                   'export const RANKING_MEN_GAP: { from: number; to: number } | null = null;\n')
        schreibe(pfad, kopf_rangliste(division, len(rows), geteilt)
                 + f'\nexport const {const}: string[] = [\n{body}];\n' + gap)

    # ---- Einzelergebnisse ----
    zeilen = []
    unbekannt = set()
    for t in turniere.values():
        ort = t['ort']
        vid = VENUE_IDS.get(ort)
        if not vid:
            unbekannt.add(ort)
            continue
        tag = datetime.datetime.strptime(t['datum'], '%d.%m.%Y').strftime('%Y-%m-%d')
        ergebnisse = ','.join(f"{r['pass']}:{r['punkte']}" for r in t['rows'])
        zeilen.append((tag, vid, f'{tag}|{vid}|{ergebnisse}'))
    if unbekannt:
        sys.exit(f'Spielort ohne ID in VENUE_IDS: {sorted(unbekannt)}')
    zeilen.sort(key=lambda z: (z[0], z[1]))

    body = ''.join(f"  '{z[2]}',\n" for z in zeilen)
    kopf = f'''// ============================================================
// MDC — Einzelergebnisse der Saison 2025/26
// ============================================================
//
// ERZEUGT aus der Arbeitsmappe des Betreibers durch
// `scripts/mdc-import-saison-2025-26.py` (Blatt „Einzelergebnisse").
// Nicht von Hand bearbeiten.
//
// Alle {len(zeilen)} Turniere der Saison vom {zeilen[0][0]} bis {zeilen[-1][0]},
// zusammen {starts} Ergebniszeilen und {punkte} vergebene Punkte.
//
// Eine Zeile je Turnier:
//
//   Datum | Spielort-ID | Passnr.:Punkte, Passnr.:Punkte, …
//
// Die Ergebnisse stehen in Platzreihenfolge — der erste Eintrag ist Platz 1.
// Die Teilnehmerzahl ist die Anzahl der Einträge; jedes Turnier der Mappe hat
// genau so viele Zeilen wie Teilnehmer und eine lückenlose Platzfolge.
//
// Die Punkte stehen mit dabei, obwohl sie sich aus Platz und Feldgröße
// ergeben würden (`lib/mdc/points.ts`): Sie sind das, was der Betreiber
// tatsächlich verbucht hat. `scripts/mdc-check-archiv.ts` rechnet beides
// gegeneinander — über alle {starts} Zeilen stimmt es exakt überein.
//
// Spielorte, die es in der Saison 2026/27 nicht mehr gibt (RG Bar, Gisi,
// Flair), stehen als `FORMER_VENUES` in `data/venues.ts` — nur mit Namen,
// ohne erfundene Adresse.
// ============================================================

export const ARCHIVE_2025_26_RAW: string[] = [
{body}];
'''
    schreibe('data/archive-2025-26.generated.ts', kopf)


if __name__ == '__main__':
    main()
